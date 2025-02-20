####################################################################################
#                                特征提取工具组                                     #
#  -- 用于训练测试数据预处理，特征工程等                                              #
#  -- 通过 import classification_tool as cft 使用 cft.func                          #
####################################################################################

#特征工程 - 文本分词需要使用
import spacy

#读取xlsx文件需要使用
from openpyxl import load_workbook

#读取csv文件需要使用
import csv
import numpy as np

#降维等其他矩阵运算
from sklearn.manifold import TSNE

class DataFeature:

    #从项目根路径开始
    dataSetPath:str = "data/"

    
################################################################
#   update on 2023-3-18                                         #
#   数据集train2.csv的排列发生了改动                              #
#   请使用excel对数据集进行一些微小的改动已支持dataFeatureList的变动#
#   使用excel打开train2.csv 选中标签所在列——排序——升序——保存改动    #
##################################################################
    
    dataFeatureList:list = \
    [
        {"fileName":"train2.csv","type":"信贷理财","range":range(2,723)},
        {"fileName":"train2.csv","type":"刷单诈骗","range":range(725,1027)},
        {"fileName":"train2.csv","type":"婚恋交友","range":range(1167,7210)},
        {"fileName":"train3.xlsx","type":"购物消费","range":range(243,430)},
        {"fileName":"train2.csv","type":"刷单诈骗","range":range(725,743)},
        {"fileName":"train2.csv","type":"婚恋交友","range":range(867,977)}
    ]
    # [
    #     {"fileName":"train2.csv","type":"信贷理财","range":range(1,515)},
    #     {"fileName":"train2.csv","type":"刷单诈骗","range":range(757,865)},
    #     {"fileName":"train2.csv","type":"婚恋交友","range":range(2150,2309)},
    #     {"fileName":"train3.xlsx","type":"购物消费","range":range(243,430)},
    #     {"fileName":"train2.csv","type":"刷单诈骗","range":range(725,743)},
    #     {"fileName":"train2.csv","type":"婚恋交友","range":range(867,977)}
    # ]


def read_csv_context(filename:str,row_range:range,col:int = 1) -> list:
    """读取csv文件并依据指定的行数范围和列数将文本装载在列表中返回
        注意由于数据集的csv文件网页快照文本往往在第2列，
        col参数默认为第二列

    Args:
        filename (str): csv文件的路径
        row_range (range): 行数范围
        col (int): 列数，单个数据

    Returns:
        list: 装载文本的列表

    Example:
        read_csv_context("xx.csv",range(1,4),1)
        return:csv文件的第2,3,4排的第2列将作为数组的元素
        ["row2 col2","row3 col2","row4 col2"]

    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    
    """
    data = []
    cnt = 0
    with open(filename,encoding="UTF8") as csvfile:
        csv_reader = csv.reader(csvfile)
        for row in csv_reader:
            if cnt in row_range:
                data.append(row[col])
            cnt += 1
    return data

def read_xlsx_context(filename:str,row_range:range,col:int) -> list:
    """读取csv文件并依据指定的行数范围和列数将文本装载在列表中返回

    Args:
        filename (str): csv文件的路径
        row_range (range):  行数范围
        col (int): 列数,由于数据集的数据排布是纵向，该方法依据一个范围的排读取指定列

    Returns:
        list: 装载文本的列表

    Example:
        read_xlsx_context("xx.xlsx",range(1,4),1)
        return:csv文件的第2,3,4排的第2列将作为数组的元素
        ["row2 col2","row3 col2","row4 col2"]
    
    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    
    """

    mapping_dict = {0:'A',1:'B',2:'C'}
    ws = load_workbook(filename)
    data:list = []

    cnt:int = 0
    for cell in ws["Sheet1"][mapping_dict[col]]:
        if cnt in row_range:
            data.append(cell.value)
        cnt += 1

    return data


def list_2_ndarray(int_lst:list) -> np.ndarray:
    """将整数列表转为numpy数组

    Args:
        int_lst (list): 整数列表

    Returns:
        np.ndarray: numpy数组

    Info:
        Created by LGD on 2023-3-19
    """
    return np.array(int_lst)

def split_word(nlp:spacy.Language,sentence:str,split_word:str = ' ') -> str:
    """将一个句子以split_word为间隔进行分词

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence (str): 未分词的中文
        split_word (str, optional): 分割词符，默认为空格. Defaults to ' '.

    Returns:
        str: 分词后的字符串

    Example:
        split_word(nlp,"唤醒自己沉睡的财产房产")
        return:"唤醒 自己 沉睡 的 财产 房产"

    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    """
    words = nlp(sentence)
    lst = [str(word) for word in words]
    #word 的数据类型为 <class 'spacy.tokens.token.Token'>

    return split_word.join(word for word in lst)


def split_word_sentence_to_split_word_list(sentence:str) -> list:
    """将分词的句子转为单词的列表

    Args:
        sentence (str): 分词的字符串

    Returns:
        list: 该字符串所有单词的列表

    Example:
        split_word_sentence_to_split_word_list("唤醒 自己 沉睡 的 财产 房产")
        return:["唤醒" ,"自己" ,"沉睡" ,"的" ,"财产" ,"房产"]

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15
    """
    return sentence.split(' ')


def split_word_arr(nlp:spacy.Language,sentence_arr:list,split_word:str = ' ') -> list:
    """将一个句子以split_word为间隔进行分词

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence (list): 未分词的中文
        split_word (str, optional): 分割词符，默认为空格. Defaults to ' '.

    Returns:
        str: 分词后的字符串

    Example:
        split_word_arr(nlp,["唤醒自己沉睡的财产房产","骗子竟是东方基金"])
        return:["唤醒 自己 沉睡 的 财产 房产","骗子 竟是 东方 基金"]

    Info:
        Created by LGD on 2023-3-9
        Last update on 2023-3-9
    """
    sentence_list = []
    for sentence in sentence_arr:
        words = nlp(sentence)
        single_sentence_lst = [str(word) for word in words]
        sentence_list.append(split_word.join(word for word in single_sentence_lst))
        
    return sentence_list

def split_word_single_arr(nlp:spacy.Language,sentence_arr:list,throwStopWord:bool = False) -> list:
    """将一个句子以单词为单位分割为单词的列表

    Args:
        nlp (spacy.Language): spacy Language模型
        sentence_arr (list): 未分词的中文

    Returns:
        list: 分词后的列表

    Example:
        split_word_arr(nlp,["唤醒自己沉睡的财产房产","骗子竟是东方基金"])
        return:[["唤醒" ,"自己" ,"沉睡", "的" ,"财产" ,"房产"],["骗子","竟是","东方" ,"基金"]]

    Info:
        Created by LGD on 2023-3-16
        Last update on 2023-3-16
    """
    word_gather_list = []
    for sentence in sentence_arr:
        words = nlp(sentence)
        word_gather_list.append([str(word) for word in words])
    
    return word_gather_list


def normalization_word_number(wordSetList:list,labelList:list,number:int) ->list:
    """归一化单词数量，当数量小于number时会被剔除，大于number时会归一到number的数量

    Args:
        wordList (list): 需要归一化的单词集，为单词集合的列表
        labelList (list):根据单词集的剔除情况

    Returns:
        list: 归一化后的列表

    Info:
        Created by LGD on 2023-3-19
    """
    i:int = 0
    cnt:int = 0
    ori_len:int = len(wordSetList)
    while i < ori_len - cnt:
        if len(wordSetList[i]) < number:
            wordSetList.pop(i)
            labelList.pop(i)
            cnt += 1
            i -= 1
        i+=1

    wordSetList = [word_set[0:600] for word_set in wordSetList]
    return wordSetList,labelList

def text_list_throw_stop_word(word_gather_list:list,stopWord:list) ->list:
    """依据停用词列表去除停用词

    Args:
        sentence_arr (list): 单词集列表
        stopWord (list): 停用词列表

    Returns:
        list: 

    Example:
        func([['天气','好'],['你','好']],['好'])
        return:[['天气'],['你']]
    """
    new_word_gather_list = []
    for word_gather in word_gather_list:
        new_word_gather = [word for word in word_gather if word not in stopWord]
        new_word_gather_list.append(new_word_gather)

    return new_word_gather_list



def wordGatherList_to_VectorList(nlp:spacy.Language,wordGatherList:list,is_flat:bool=False) -> list:
    """将单词集合的列表转换为词向量的列表

    Args:
        nlp (spacy.Language): spacy Language模型
        wordGatherList (list): 单词集合列表

    Returns:
        list: 词向量的列表

    Example:
        wordGatherList_to_VectorList([['cat','dog'],['red','blue']])
        result: [[333,222,444],[555,333,444]]  每个元素为输入的每个单词集合的词向量的拼接

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15

    Update:
        2023-3-16 将输出结果由词向量的扁平拼接改为（单词数,300)的矩阵
    """
    embedding = []
    embedding_list = []
    for word_gather in wordGatherList:
        for word in word_gather:
            embedding:np.ndarray = np.append(embedding,nlp.vocab[word].vector)
        embedding_list.append(embedding)
        embedding = []

    new_embedding_list = []
    for i in range(len(wordGatherList)):
        embedding = embedding_list[i].reshape(len(wordGatherList[i]),-1)
        new_embedding_list.append(embedding)
        

    return new_embedding_list

def wordGatherList_to_Matrix(nlp:spacy.Language,wordGatherList:list,is_flat:bool=False) -> np.ndarray:
    """将单词集合的列表转换为词向量拼接的二维矩阵
        注意，由于python列表的元素长度可以任意，
            但二维ndarray的列数不可以，要求wordGather的长度必须归一化

    Args:
        nlp (spacy.Language): spacy Language模型
        wordGatherList (list): 单词集合列表

    Returns:
        list: 词向量的列表

    Example:
        wordGatherList_to_Matrix([['cat','dog'],['red','blue']])
        result: [[333,222,444],[555,333,444]]  每个元素为输入的每个单词集合的词向量的拼接

    Info:
        Created by LGD on 2023-3-15
        Last update on 2023-3-15

    Update:
        2023-3-16 将输出结果由词向量的扁平拼接改为（单词数,300)的矩阵
    """
    embedding = []
    for word_gather in wordGatherList:
        for word in word_gather:
            embedding:np.ndarray = np.append(embedding,nlp.vocab[word].vector)

    embedding:np.ndarray = embedding.reshape(len(wordGatherList),-1)

    if(not is_flat):
        embedding = embedding.reshape(len(wordGatherList),len(wordGatherList[0]),300)

    return embedding


def lower_dimension(array:np.ndarray) -> np.ndarray:
    """对高维向量进行降维
       似乎可以将任一的高维矩阵降成2维        
    
    Args:
        array (np.ndarray): 高维向量

    Returns:
        np.ndarray: 降维向量

    Info:
        Created by LGD on 2023-3-16
        Last update on 2023-3-16
    """
    tsne = TSNE()
    return tsne.fit_transform(array)


#TODO 尚未完成
def count_word_freq(word_gather:list) -> dict:
    """获取一个单词的集合（列表），将返回一个记录了所有单词的出现次数的字典

    Args:
        word_gather (list): 单词集合

    Returns:
        dict: 词频字典

    Example:
        count_word_freq(["cat","cat","mouse"])
        return:{"cat":2,"mouse":1}

    Info:
        Created by LGD on 2023-3-16
        Last update on 2023-3-16
    """
    pass

